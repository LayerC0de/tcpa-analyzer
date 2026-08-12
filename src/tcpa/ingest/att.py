"""Ingest AT&T "Usage Details" .xlsx exports.

AT&T records are BROADER but SHALLOWER than the device call log, and mixing the
two naively corrupts the analysis. The differences that matter:

  TIME      ~16 bill cycles vs the phone's ~2000-entry cap (5 months here).
            This is the only reason to use AT&T data at all.

  DURATION  Billed MINUTES, rounded up -- a 9-second call reads as "1".
            The device log stores exact seconds. Any analysis keying on precise
            duration (the burn-after-use fingerprint, sub-10-second drops) MUST
            run on source='android' rows only. Rows from here carry
            duration_estimated=1 so that filter is possible.

  MISSED    Carriers bill usage, so calls that rang without being answered do
            not appear at all. The campaign's signature is one answered call
            PLUS one zero-duration call; AT&T can only ever show the first.
            Expect roughly half the campaign's events to be invisible here.

  GEO       Incoming rows carry the literal string "Incoming" in the Location
            column rather than a city, so no geography for inbound calls.

  TEXTS     The Text sheet is a bonus: texts to a cell are "calls" under the
            TCPA, so they are a separate violation surface the phone log
            export did not give us.
"""
from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path

from ..phone import normalize

HEADER_ROW_TEXT = "Incoming/Outgoing"
CYCLE_RE = re.compile(r"Bill Cycle:\s*(.+)", re.I)
PHONE_RE = re.compile(r"Phone Number:\s*([\d.\-() ]+)", re.I)


def _local(date_str: str, time_str: str, tz_name: str = "America/New_York"):
    """AT&T prints 'Aug 12, 2026' / '5:21 pm' in the account's local time."""
    stamp = f"{date_str.strip()} {time_str.strip().upper()}"
    dt = datetime.strptime(stamp, "%b %d, %Y %I:%M %p")
    try:
        from zoneinfo import ZoneInfo
        return dt.replace(tzinfo=ZoneInfo(tz_name))
    except Exception:
        return dt.replace(tzinfo=timezone.utc)


def _rows(ws):
    """Yield dict rows from a sheet, skipping the 5-line AT&T preamble."""
    header = None
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if not any(cells):
            continue
        if header is None:
            if cells[0] == HEADER_ROW_TEXT or cells[0] == "Date":
                header = cells
            continue
        yield dict(zip(header, cells))


def parse(path: Path, tz_name: str = "America/New_York") -> dict:
    """Return {'calls': [...], 'texts': [...], 'cycle': str, 'line': str}."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    cycle, line = None, None
    ws = wb["Talk"] if "Talk" in wb.sheetnames else wb.worksheets[0]
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        text = str(row[0] or "")
        if m := CYCLE_RE.search(text):
            cycle = m.group(1).strip()
        if m := PHONE_RE.search(text):
            line = normalize(m.group(1))

    calls = []
    for r in _rows(ws):
        number = normalize(r.get("Contact", ""))
        try:
            local = _local(r["Date"], r["Time"], tz_name)
        except (KeyError, ValueError):
            continue
        minutes = r.get("Minutes", "0")
        try:
            secs = int(float(minutes)) * 60
        except ValueError:
            secs = 0
        calls.append({
            "source": "att",
            # Composite key: the export carries no row id, so identity is the
            # line + timestamp + counterparty. Re-importing the same cycle is
            # then idempotent rather than duplicating every row.
            "source_row_id": f"{line}|{local.isoformat()}|{r.get('Contact','')}",
            "number_raw": r.get("Contact"),
            "number": number,
            "contact_name": None,
            "ts_utc": int(local.timestamp() * 1000),
            "local_iso": local.strftime("%Y-%m-%d %H:%M:%S"),
            "local_date": local.strftime("%Y-%m-%d"),
            "local_hour": local.hour,
            "duration_s": secs,
            "direction": (r.get("Incoming/Outgoing") or "").upper() or "UNKNOWN",
            "presentation": None,
            "block_reason": 0,
            # "Incoming" is a placeholder, not a place.
            "geo": None if (r.get("Location") or "").lower() == "incoming" else r.get("Location"),
            "own_number": line,
            "transcription": None,
        })

    texts = []
    if "Text" in wb.sheetnames:
        for r in _rows(wb["Text"]):
            number = normalize(r.get("Contact", ""))
            try:
                local = _local(r["Date"], r["Time"], tz_name)
            except (KeyError, ValueError):
                continue
            texts.append({
                "number": number, "number_raw": r.get("Contact"),
                "direction": (r.get("Incoming/Outgoing") or "").upper(),
                "kind": r.get("Type"), "ts_utc": int(local.timestamp() * 1000),
                "local_iso": local.strftime("%Y-%m-%d %H:%M:%S"),
                "local_date": local.strftime("%Y-%m-%d"),
                "own_number": line,
            })

    wb.close()
    return {"calls": calls, "texts": texts, "cycle": cycle, "line": line}


def load_texts(con, texts: list[dict]) -> int:
    """Store text records. Identity is line + timestamp + counterparty."""
    before = con.execute("SELECT COUNT(*) FROM texts").fetchone()[0]
    con.executemany("""
        INSERT OR IGNORE INTO texts
          (source, source_row_id, number, number_raw, direction, kind,
           ts_utc, local_iso, local_date, own_number)
        VALUES ('att',?,?,?,?,?,?,?,?,?)
    """, [(f"{t['own_number']}|{t['local_iso']}|{t['number_raw']}|{t['kind']}",
           t["number"], t["number_raw"], t["direction"], t["kind"],
           t["ts_utc"], t["local_iso"], t["local_date"], t["own_number"])
          for t in texts])
    con.commit()
    return con.execute("SELECT COUNT(*) FROM texts").fetchone()[0] - before


DUP_WINDOW_MS = 120_000  # AT&T rounds to the minute; device time is exact


def load(con, calls: list[dict]) -> dict:
    """Insert AT&T calls, flagging any that duplicate a device-log record.

    The two sources overlap wherever the phone log still reaches. The same call
    must not be counted twice, and where both exist the DEVICE row wins because
    its duration is exact. Duplicates are kept rather than dropped so the two
    sources can still be compared -- that comparison is what reveals how much
    the carrier omits.
    """
    from ..ingest.android import COLUMNS

    stats = {"inserted": 0, "duplicates": 0}
    for c in calls:
        dup = con.execute("""
            SELECT 1 FROM calls
            WHERE source = 'android' AND number IS ?
              AND ABS(ts_utc - ?) <= ?
            LIMIT 1
        """, (c["number"], c["ts_utc"], DUP_WINDOW_MS)).fetchone()
        cur = con.execute(
            f"INSERT OR IGNORE INTO calls ({','.join(COLUMNS)}, duration_estimated, dup_of_device) "
            f"VALUES ({','.join('?' * len(COLUMNS))}, 1, ?)",
            tuple(c[k] for k in COLUMNS) + (1 if dup else 0,),
        )
        if cur.rowcount:
            stats["inserted"] += 1
            if dup:
                stats["duplicates"] += 1
    con.commit()
    return stats
