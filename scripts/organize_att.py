#!/usr/bin/env python3
"""Rename AT&T exports by their INTERNAL bill-cycle label and report gaps.

Every download arrives as `<line>_UsageDetails.xlsx`, so Chrome disambiguates
with (1), (2)... which carries no meaning. The authoritative cycle is printed
inside the file, so identity is read from the content rather than the filename.
That makes re-running safe and makes duplicate downloads self-evident.

Run:  python scripts/organize_att.py
"""
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"

CYCLE_RE = re.compile(r"Bill Cycle:\s*(.+)", re.I)

# The full set offered by AT&T, newest first.
EXPECTED = [
    "Aug 02, 2026 - Sep 01, 2026", "Jul 02, 2026 - Aug 01, 2026",
    "Jun 02, 2026 - Jul 01, 2026", "May 02, 2026 - Jun 01, 2026",
    "Apr 02, 2026 - May 01, 2026", "Mar 02, 2026 - Apr 01, 2026",
    "Feb 02, 2026 - Mar 01, 2026", "Jan 02, 2026 - Feb 01, 2026",
    "Dec 02, 2025 - Jan 01, 2026", "Nov 02, 2025 - Dec 01, 2025",
    "Oct 02, 2025 - Nov 01, 2025", "Sep 02, 2025 - Oct 01, 2025",
    "Aug 02, 2025 - Sep 01, 2025", "Jul 02, 2025 - Aug 01, 2025",
    "Jun 02, 2025 - Jul 01, 2025", "May 02, 2025 - Jun 01, 2025",
    "Apr 02, 2025 - May 01, 2025",
]


def cycle_of(path: Path) -> str | None:
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    ws = wb["Talk"] if "Talk" in wb.sheetnames else wb.worksheets[0]
    label = None
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        if row and row[0] and (m := CYCLE_RE.search(str(row[0]))):
            label = m.group(1).strip()
            break
    wb.close()
    return label


def slug(cycle: str) -> str:
    """'Apr 02, 2025 - May 01, 2025' -> 'att_2025-04'."""
    start = cycle.split("-")[0].strip()
    return "att_" + datetime.strptime(start, "%b %d, %Y").strftime("%Y-%m")


def main():
    RAW.mkdir(parents=True, exist_ok=True)
    loose = sorted(ROOT.glob("*.xlsx")) + sorted(RAW.glob("*.xlsx"))
    if not loose:
        sys.exit("no .xlsx files found")

    seen: dict[str, Path] = {}
    dupes, unreadable = [], []

    for path in loose:
        cycle = cycle_of(path)
        if not cycle:
            unreadable.append(path.name)
            continue
        target = RAW / f"{slug(cycle)}.xlsx"
        if cycle in seen:
            dupes.append((path.name, cycle))
            if path.resolve() != seen[cycle].resolve():
                path.unlink()
            continue
        if path.resolve() != target.resolve():
            shutil.move(str(path), str(target))
        seen[cycle] = target
        print(f"  {target.name:<20} <- {cycle}")

    print(f"\nkept {len(seen)} distinct cycles")
    if dupes:
        print(f"removed {len(dupes)} duplicate download(s): "
              + ", ".join(f"{n} ({c.split('-')[0].strip()})" for n, c in dupes))
    if unreadable:
        print(f"unreadable: {unreadable}")

    missing = [c for c in EXPECTED if c not in seen]
    print(f"\n{len(EXPECTED) - len(missing)}/{len(EXPECTED)} cycles present")
    if missing:
        print("MISSING:")
        for c in missing:
            print(f"  {c}")
        print("\nJS to re-pull just these (paste labels into window.__pull):")
        print("  " + str(missing))
    else:
        print("complete -- all cycles downloaded")


if __name__ == "__main__":
    main()
